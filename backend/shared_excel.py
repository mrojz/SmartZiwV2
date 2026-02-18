"""
Shared Excel output and keyword configuration for all scrapers.
"""

import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── Keywords shared by all scrapers ──────────────────────────────────────────

_DEFAULT_KEYWORDS = [
    "Cybersecurity",
    "Cyber Security",
    "Cybersécurité",
    "Penetration Testing",
    "Pentest",
    "Ethical Hacking",
    "tests d'intrusion",
    "Vulnerability Assessment",
    "VAPT",
    "IT Security",
    "OT Security",
    "ISMS",
    "SMSI",
    "ISO 27001",
    "PCI DSS",
    "Phishing",
    "social engineering",
    "ingénieurie sociale",
    "SWIFT CSP",
    "CSCF",
    "DORA",
    "CISO assistance",
    "assistance RSSI",
]


def _load_keywords():
    """Load keywords from config.json, fallback to defaults."""
    import json as _json
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            cfg = _json.load(f)
        kw = cfg.get("keywords", [])
        return kw if kw else _DEFAULT_KEYWORDS
    except (FileNotFoundError, _json.JSONDecodeError):
        return _DEFAULT_KEYWORDS


SEARCH_KEYWORDS = _load_keywords()


# ── Date formatting helper ───────────────────────────────────────────────────

def format_date(value):
    """Convert various date formats to MM/DD/YYYY (e.g. '10/21/2024').
    Handles: Unix timestamps (ms or s), ISO strings, Power BI serial day offsets,
    and World Bank date strings like '04-Feb-2026'."""
    if not value:
        return ""
    try:
        val = str(value).strip()
        # Try common formats
        for fmt in (
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%dT%H:%M:%SZ",
            "%Y-%m-%d",
            "%d-%b-%Y",       # World Bank: 04-Feb-2026
            "%d-%B-%Y",       # 04-February-2026
            "%m/%d/%Y",       # Already in target format
            "%m/%d/%Y %I:%M:%S %p",  # World Bank closing date format
        ):
            try:
                return datetime.strptime(val, fmt).strftime("%m/%d/%Y")
            except ValueError:
                continue
        # Numeric: Unix ms, Unix s, or serial day offset
        num = float(val)
        if num > 1e12:
            return datetime.utcfromtimestamp(num / 1000).strftime("%m/%d/%Y")
        elif num > 1e9:
            return datetime.utcfromtimestamp(num).strftime("%m/%d/%Y")
        elif num > 0:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(num))).strftime("%m/%d/%Y")
    except Exception:
        pass
    return str(value)


def parse_date(date_str):
    """Parse MM/DD/YYYY string to datetime. Returns None on failure."""
    try:
        return datetime.strptime(date_str, "%m/%d/%Y")
    except (ValueError, TypeError):
        return None


def is_expired(project):
    """Return True if the project's due date has already passed."""
    due = project.get("project_end_date", "")
    if not due:
        return False  # No due date → keep it
    dt = parse_date(due)
    if dt is None:
        return False
    return dt < datetime.now()


# ── Excel columns ────────────────────────────────────────────────────────────

EXCEL_COLUMNS = [
    ("project_id",          "Project ID"),
    ("project_name",        "Project Name"),
    ("project_start_date",  "Publication Date"),
    ("project_end_date",    "Due Date"),
    ("project_description", "Notice Title"),
    ("project_sponsor",     "Country"),
    ("source",              "Source"),
    ("document_url",        "Document URL"),
    ("project_url",         "Project URL"),
    ("matched_keywords",    "Matched Keywords"),
    ("ai_verified",         "AI Verified"),
    ("decision",            "Decision"),
]


def load_existing_projects(filename="projects.xlsx"):
    """Load ALL existing project rows from the Excel file.
    
    Returns:
        existing_keys: set of (project_id, notice_title) for quick lookup
        existing_rows: list of dicts (one per row) with all column values preserved
    """
    existing_keys = set()
    existing_rows = []

    if not os.path.exists(filename):
        return existing_keys, existing_rows

    try:
        from openpyxl import load_workbook
        old_wb = load_workbook(filename)
        old_ws = old_wb.active

        # Map header names to column indices
        header_map = {}
        for col_idx in range(1, old_ws.max_column + 1):
            val = old_ws.cell(row=1, column=col_idx).value
            if val:
                header_map[val] = col_idx

        # Reverse map: col_idx -> dict_key
        label_to_key = {label: key for key, label in EXCEL_COLUMNS}

        for row in range(2, old_ws.max_row + 1):
            row_dict = {}
            for label, col_idx in header_map.items():
                dict_key = label_to_key.get(label)
                if dict_key:
                    row_dict[dict_key] = old_ws.cell(row=row, column=col_idx).value or ""
            
            if row_dict:
                pid = str(row_dict.get("project_id", ""))
                title = str(row_dict.get("project_description", ""))
                existing_keys.add((pid, title))
                existing_rows.append(row_dict)

        old_wb.close()
        print(f"[+] Loaded {len(existing_rows)} existing projects from '{filename}'")
    except Exception as e:
        print(f"[!] Could not load existing file: {e}")

    return existing_keys, existing_rows


def save_to_excel(all_rows, filename="projects.xlsx"):
    """Write ALL rows (old + new) to a styled Excel file.
    
    This is called with the merged list: existing rows (untouched) + new rows (appended).
    Rows are sorted by Publication Date (newest first).
    """
    if not all_rows:
        print("[!] No projects to save")
        return

    columns = EXCEL_COLUMNS

    # Sort by publication date (newest first)
    all_rows.sort(
        key=lambda p: parse_date(p.get("project_start_date", "")) or datetime.min,
        reverse=True,
    )

    # ── Build workbook ──
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    rejected_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    # Header row
    for col_idx, (key, label) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 30

    # Data rows
    data_align = Alignment(vertical="center", wrap_text=False)
    for row_idx, project in enumerate(all_rows, 2):
        is_rejected = project.get("ai_verified", "") == "No"
        for col_idx, (key, _) in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=project.get(key, ""))
            cell.alignment = data_align
            cell.border = thin_border
            if is_rejected:
                cell.fill = rejected_fill

    # Auto-adjust column widths
    for col_idx, (key, label) in enumerate(columns, 1):
        max_length = len(label)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 55)

    # Auto-filter
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{len(all_rows) + 1}"

    # Freeze header
    ws.freeze_panes = "A2"

    # Decision dropdown validation (last column)
    dec_col = get_column_letter(len(columns))
    dv = DataValidation(
        type="list",
        formula1='"Go,No Go"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Please select Go or No Go"
    dv.prompt = "Choose Go or No Go"
    dv.promptTitle = "Decision"
    dv.sqref = f"{dec_col}2:{dec_col}{len(all_rows) + 1}"
    ws.add_data_validation(dv)

    wb.save(filename)
    print(f"[+] Saved {len(all_rows)} projects to '{filename}'")
